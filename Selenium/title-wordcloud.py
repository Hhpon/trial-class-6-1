from openpyxl import load_workbook
from wordcloud import WordCloud
import jieba

print('请输入需要生成词云的excel文件')
file_name = input('')


def get_row_value(sheet, column):
    row_value = []
    for i in range(2, sheet.max_row+1):
        row_value.append(sheet.cell(row=i, column=column).value)
    return row_value


stopworld = ('这', '那', '你', '我', '他', '她', '它')


def gen_wordcloud(data, font, out_img, width, height):
    word_str = ''.join(data)
    jieba_str = jieba.lcut(word_str)
    cut_word = ''
    for i in jieba_str:
        if i not in stopworld:
            cut_word += i
    wc = WordCloud(width=width, height=height,
                   background_color='white', font_path=font, mask=None)
    print(cut_word)
    wc.generate(word_str)
    wc.to_file(out_img)


workbook = load_workbook('Selenium/%s.xlsx' % file_name)
sheetnames = workbook.sheetnames
sheet = workbook[sheetnames[0]]
row_value = get_row_value(sheet, 1)
gen_wordcloud(row_value, 'Selenium/FZNSTJW.TTF',
              'Selenium/out_img_test.png', 1800, 1500)
