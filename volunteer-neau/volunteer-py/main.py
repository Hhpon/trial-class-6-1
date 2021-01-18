from openpyxl import load_workbook
from openpyxl import Workbook
import os
import pymongo
import time

client = pymongo.MongoClient(host='localhost')
db = client['trial-class']
collection = db['volunteers']


def read_volunteer(root_dir, first_level_dir, file):
    workbook = load_workbook('%s/%s/%s' % (root_dir, first_level_dir, file),
                             data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    cells = list(sheet[sheet.dimensions])

    keys = [
        'num', 'studentName', 'factory', 'classroom', 'studentId',
        'totalServiceTime', 'org', 'service', 'duration'
    ]
    volunteers = []

    for row in cells[1:]:
        if row[0].value:
            values = list(map(lambda cell: cell.value, row))
            volunteer_info = dict(zip(keys, values))
        elif row[7].value:
            values = list(
                value for value in volunteers[len(volunteers) - 1].values())
            volunteer_info = dict(zip(keys, values))
            volunteer_info['service'] = row[7].value
            volunteer_info['duration'] = row[8].value
        volunteers.append(volunteer_info)
    return volunteers


def save_volunteer(volunteer_info):
    if isinstance(volunteer_info, list):
        for volunteer in volunteer_info:
            print(volunteer)
            collection.insert_one({
                'studentName': volunteer['studentName'],
                'factory': volunteer['factory'],
                'classroom': volunteer['classroom'],
                'studentId': volunteer['studentId'],
                'org': volunteer['org'],
                'service': volunteer['service'],
                'duration': volunteer['duration']
            })
    elif isinstance(volunteer_info, dict):
        collection.insert_one({
            'studentName': volunteer_info['studentName'],
            'factory': volunteer_info['factory'],
            'classroom': volunteer_info['classroom'],
            'studentId': volunteer_info['studentId'],
            'org': volunteer_info['org'],
            'service': volunteer_info['service'],
            'duration': volunteer_info['duration']
        })


def main():
    root_dir = 'volunteer-neau/volunteer-py/volunteer-info'
    first_level_dirs = os.listdir(root_dir)
    print(first_level_dirs)
    for first_level_dir in first_level_dirs:
        files = os.listdir('%s/%s' % (root_dir, first_level_dir))
        print(files)
        for file in files:
            print('%s/%s/%s' % (root_dir, first_level_dir, file))
            volunteers = read_volunteer(root_dir, first_level_dir, file)
            save_volunteer(volunteers)


if __name__ == "__main__":
    main()