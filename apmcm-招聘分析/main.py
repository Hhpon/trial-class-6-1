from openpyxl import load_workbook
from openpyxl import Workbook
import os

month_dict = {
    'Jan': '01',
    'Feb': '02',
    'Mar': '03',
    'Apr': '04',
    'May': '05',
    'June': '06',
    'July': '07',
    'Aug': '08',
    'Sept': '09',
    'Oct': '10',
    'Nov': '11',
    'Dec': '12'
}


def get_sector_edu():
    workbook = load_workbook('apmcm-招聘分析/market-demand/2015/09.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    return {
        'sectors': list(sector[0].value for sector in sheet['B4:B50']),
        'edus': list(edu.value for edu in sheet['D3:L3'][0])
    }


def read_demand(dir, first_level_dir, file):
    workbook = load_workbook('%s/%s/%s' % (dir, first_level_dir, file))
    sheet = workbook[workbook.sheetnames[0]]
    cells = list(sheet[sheet.dimensions])
    demand_lists = []
    keys = list(
        map(lambda cell: cell.value, cells[1:3][0][:3] + cells[1:3][1][3:]))
    keys.extend(['month', 'year'])
    for row in cells[3:]:
        row_values = list(map(lambda cell: cell.value, row))
        row_values.extend([file.split('.')[0], first_level_dir])
        demand = dict(zip(keys, row_values))
        demand_lists.append(demand)
    return demand_lists


def write_demand(sectors, demand_lists, edu, months, dir):
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = 'sheet1'
    first_row = months[:]
    first_row.insert(0, '')
    print(first_row)
    sheet1.append(first_row)
    for sector in sectors:
        row = [sector]
        for month in months:
            current_total_demand = 0
            for demand in filter(
                    lambda demand: demand.get('Sector') == sector and demand.
                    get('month') == month_dict.get(month), demand_lists):
                current_total_demand += demand[edu]
            row.append(current_total_demand)
        print(row)
        sheet1.append(row)
    workbook.save(filename='%s/%s.xlsx' % (dir, edu))


def main():
    demand_dir = 'apmcm-招聘分析/market-demand'
    demand_paths = os.listdir(demand_dir)
    demand_lists = []
    for demand_path in demand_paths:
        print(demand_path)
        if demand_path == '.DS_Store':
            continue
        files = os.listdir('%s/%s' % (demand_dir, demand_path))
        for file in files:
            if file == '.DS_Store':
                continue
            demand_lists.extend(read_demand(demand_dir, demand_path, file))
    recruit_info = get_sector_edu()
    sector_dir = 'apmcm-招聘分析/market-sector'
    months = [
        'Jan', 'Feb', 'Mar', 'Apr', 'May', 'June', 'July', 'Aug', 'Sept',
        'Oct', 'Nov', 'Dec'
    ]
    for edu in recruit_info['edus']:
        write_demand(recruit_info['sectors'], demand_lists, edu, months,
                     sector_dir)


if __name__ == "__main__":
    main()
